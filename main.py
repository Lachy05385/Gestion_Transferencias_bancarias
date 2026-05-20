from fastapi import FastAPI, Depends, HTTPException, status
#from app.routers.admin_usuarios import router
from sqlalchemy.orm import joinedload
from sqlalchemy.orm import joinedload      # Para eager loading (cargar relaciones)
from sqlalchemy.orm import selectinload    # Alternativa a joinedload (mejor para relaciones uno a muchos)
from sqlalchemy.orm import subqueryload    # Otra alternativa
from sqlalchemy.orm import lazyload        # Para carga perezosa (default)
from sqlalchemy.orm import contains_eager
import re



from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from typing import List, Optional
from fastapi.security import OAuth2PasswordRequestForm
from fastapi import FastAPI, Depends, HTTPException, status, Form
from app import crud, models, auth, schemas, utils
from app.database import engine, get_db, create_tables
from pydantic import Field  # ← ESTA ES LA IMPORTACIÓN QUE FALTA
from fastapi import Query  # Importa Query directamente
from datetime import timedelta
from fastapi.responses import HTMLResponse
from fastapi import APIRouter, Request
from app.models import EstadoTransaccion
from app.auth import get_password_hash
from app.auth import verify_password


# Crear tablas
create_tables()

app = FastAPI(
    title="API de Gestión Bancaria",
    description="API para procesar y gestionar transacciones bancarias con usuarios",
    version="2.0.0"
)

# Configurar CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Endpoints públicos
# main.py - Endpoints para 

@app.get("/landing", response_class=HTMLResponse)
async def landing_page(request: Request):
    return templates.TemplateResponse("landing.html", {"request": request})





#EMPRESAS
@app.post("/empresas", response_model=schemas.EmpresaResponse)
def crear_empresa(
    empresa: schemas.EmpresaCreate,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """Crear una nueva empresa (solo admin)"""
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No tienes permisos")
    
    # Verificar si NIT ya existe
    existente = db.query(models.Empresa).filter(models.Empresa.nit == empresa.nit).first()
    if existente:
        raise HTTPException(status_code=400, detail="El NIT ya está registrado")
    
    nueva_empresa = models.Empresa(
        nombre=empresa.nombre,
        nit=empresa.nit,
        direccion=empresa.direccion,
        telefono=empresa.telefono
    )
    
    db.add(nueva_empresa)
    db.commit()
    db.refresh(nueva_empresa)
    return nueva_empresa

@app.get("/empresas", response_model=List[schemas.EmpresaResponse])
def listar_empresas(
    skip: int = 0,
    limit: int = 100,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """Listar empresas (solo admin)"""
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No tienes permisos")
    
    empresas = db.query(models.Empresa).offset(skip).limit(limit).all()
    return empresas

# ========== ENDPOINT PARA CONTADOR (VER TRANSACCIONES DE LA EMPRESA) ==========

@app.get("/transacciones/empresa", response_model=List[schemas.TransaccionResponse])
def listar_transacciones_empresa(
    usuario_id: Optional[int] = None,
    estado: Optional[str] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    empresa_id: Optional[int] = None,  # 👈 AGREGAR ESTE PARÁMETRO
    skip: int = 0,
    limit: int = 100,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Listar transacciones de la empresa (solo para contadores y admins)
    - Contador: ve todas las transacciones de su empresa (ignora empresa_id)
    - Admin: puede ver transacciones de cualquier empresa (usando empresa_id)
    """
    # Verificar permisos
    if current_user.rol not in ["contador", "admin"]:
        raise HTTPException(
            status_code=403, 
            detail="No tienes permisos. Solo contadores y administradores pueden ver transacciones de la empresa"
        )
    
    print(f"🔍 Usuario: {current_user.email} (rol: {current_user.rol}, empresa_id: {current_user.empresa_id})")
    print(f"🔍 Parámetros - usuario_id: {usuario_id}, empresa_id: {empresa_id}")
    
    # Construir consulta
    query = db.query(models.Transaccion).options(joinedload(models.Transaccion.usuario))
    
    # 👇 LÓGICA CORREGIDA
    if current_user.rol == "contador":
        # Contador: SOLO puede ver transacciones de su propia empresa
        query = query.filter(models.Transaccion.empresa_id == current_user.empresa_id)
        print(f"🏢 Contador filtrando por empresa_id: {current_user.empresa_id}")
    elif current_user.rol == "admin":
        # Admin: puede ver de cualquier empresa (si se especifica, sino todas)
        if empresa_id is not None:
            query = query.filter(models.Transaccion.empresa_id == empresa_id)
            print(f"👑 Admin filtrando por empresa_id: {empresa_id}")
        else:
            print(f"👑 Admin viendo todas las empresas")
    
    # Filtro por usuario (opcional)
    if usuario_id is not None:
        query = query.filter(models.Transaccion.usuario_id == usuario_id)
        print(f"👤 Filtrando por usuario_id: {usuario_id}")
    
    # Filtros adicionales
    if estado:
        query = query.filter(models.Transaccion.estado == estado)
    if fecha_inicio:
        query = query.filter(models.Transaccion.fecha >= fecha_inicio)
    if fecha_fin:
        query = query.filter(models.Transaccion.fecha <= fecha_fin)
    
    # Ordenar y ejecutar
    transacciones = query.order_by(
        models.Transaccion.fecha.desc()
    ).offset(skip).limit(limit).all()
    
    print(f"📊 Transacciones encontradas: {len(transacciones)}")
    
    return transacciones


# ========== ENDPOINT PARA OBTENER USUARIOS DE LA EMPRESA ==========

@app.get("/empresa/usuarios", response_model=List[schemas.UserResponse])
def listar_usuarios_empresa(
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Listar todos los usuarios de la empresa del contador
    (Para que el contador pueda filtrar por usuario)
    """
    if current_user.rol not in ["contador", "admin"]:
        raise HTTPException(
            status_code=403, 
            detail="No tienes permisos"
        )
    
    usuarios = db.query(models.Usuario).filter(
        models.Usuario.empresa_id == current_user.empresa_id
    ).all()
    
    return usuarios

#usuarios
@app.get("/admin/usuarios", response_model=List[schemas.UserResponse])
def listar_usuarios(
    empresa_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 100,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Listar todos los usuarios (solo administradores)
    Puede filtrar por empresa_id
    """
    # Verificar que quien consulta es admin
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No tienes permisos de administrador")
    
    # Construir consulta
    query = db.query(models.Usuario)
    
    # Filtrar por empresa si se especifica
    if empresa_id:
        query = query.filter(models.Usuario.empresa_id == empresa_id)
    
    # Ordenar y paginar
    usuarios = query.order_by(models.Usuario.id).offset(skip).limit(limit).all()
    
    return usuarios


@app.post("/registro", response_model=schemas.UserResponse)
def registrar_usuario(usuario: schemas.UserCreate, db: Session = Depends(get_db)):
    print("recibido", usuario)
    db_user = crud.get_user_by_email(db, email=usuario.email)
    if db_user:
        raise HTTPException(
            status_code=400,
            detail="El email ya está registrado"
        )
    return crud.create_user(db=db, user=usuario)

#=========  MODIFICAR USUARIOS ===================
from fastapi import Body, HTTPException, Depends

@app.patch("/usuarios/{usuario_id}", response_model=schemas.UserResponse)
def actualizar_usuario(
    usuario_id: int,
    nombre: Optional[str] = Body(None),
    apellido: Optional[str] = Body(None),
    password_actual: Optional[str] = Body(None),
    nueva_password: Optional[str] = Body(None),
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    print("Buscando al usuario con id: ", usuario_id)
    usuario = db.query(models.Usuario).filter(models.Usuario.id == usuario_id).first()
    if not usuario:
        # 🔴 Directly raise a 404 error.
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Usuario con id {usuario_id} no encontrado"
        )
    """
    Actualizar nombre, apellido o contraseña de un usuario.
    - El email NO se puede modificar.
    - Si se envía nueva_password, se requiere password_actual (a menos que el usuario sea admin).
    """
    # Verificar que el usuario existe
    print("Usuario Encontrado ", usuario)

    # Permisos: solo el mismo usuario o un admin
    if current_user.id != usuario_id and current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No tienes permisos para modificar este usuario")

    # Actualizar nombre y apellido (si se proporcionan)
    if nombre is not None:
        usuario.nombre = nombre
    if apellido is not None:
        usuario.apellido = apellido

    # Cambio de contraseña
    if nueva_password is not None:
        # Si no es admin, validar la contraseña actual
        if current_user.rol != "admin":
            if not password_actual:
                raise HTTPException(status_code=400, detail="Se requiere la contraseña actual para cambiar la contraseña")
            if not verify_password(password_actual, usuario.hashed_password):
                raise HTTPException(status_code=400, detail="Contraseña actual incorrecta")
        # Hashear nueva contraseña
        usuario.hashed_password = get_password_hash(nueva_password)

    db.commit()
    db.refresh(usuario)
    return usuario


# ========== ENDPOINTS PARA ADMIN (ELIMINAR Y EDITAR USUARIOS) ==========

@app.delete("/admin/usuarios/{usuario_id}")
def eliminar_usuario(
    usuario_id: int,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Eliminar un usuario (solo administradores)
    No se puede eliminar a sí mismo
    """
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No tienes permisos de administrador")
    
    # No permitir eliminarse a sí mismo
    if current_user.id == usuario_id:
        raise HTTPException(status_code=400, detail="No puedes eliminar tu propio usuario")
    
    # Buscar usuario
    usuario = db.query(models.Usuario).filter(models.Usuario.id == usuario_id).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # Eliminar usuario (las transacciones se eliminan en cascada si tienes cascade configurado)
    db.delete(usuario)
    db.commit()
    
    return {"message": f"Usuario {usuario.email} eliminado correctamente"}

@app.put("/admin/usuarios/{usuario_id}", response_model=schemas.UserResponse)
def editar_usuario(
    usuario_id: int,
    usuario_update: schemas.UserUpdate,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Editar un usuario (solo administradores)
    """
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No tienes permisos de administrador")
    
    usuario = db.query(models.Usuario).filter(models.Usuario.id == usuario_id).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # Actualizar campos
    if usuario_update.nombre is not None:
        usuario.nombre = usuario_update.nombre
    if usuario_update.apellido is not None:
        usuario.apellido = usuario_update.apellido
    if usuario_update.rol is not None:
        usuario.rol = usuario_update.rol
    if usuario_update.empresa_id is not None:
        empresa = db.query(models.Empresa).filter(models.Empresa.id == usuario_update.empresa_id).first()
        if not empresa:
            raise HTTPException(status_code=400, detail="La empresa no existe")
        usuario.empresa_id = usuario_update.empresa_id
    if usuario_update.esta_activo is not None:
        usuario.esta_activo = usuario_update.esta_activo
    if usuario_update.password is not None and usuario_update.password.strip():
        usuario.hashed_password = auth.get_password_hash(usuario_update.password)
    
    try:
        db.commit()
        db.refresh(usuario)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al guardar: {str(e)}")
    
    # 👇 IMPORTANTE: retornar el usuario actualizado
    return usuario
    
@app.patch("/admin/usuarios/{usuario_id}/desactivar")
def desactivar_usuario(
    usuario_id: int,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No tienes permisos")
    if current_user.id == usuario_id:
        raise HTTPException(status_code=400, detail="No puedes desactivarte a ti mismo")
    usuario = db.query(models.Usuario).filter(models.Usuario.id == usuario_id).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    usuario.esta_activo = False
    db.commit()
    return {"message": f"Usuario {usuario.email} desactivado", "usuario_id": usuario_id}

# Reactivar usuario
@app.patch("/admin/usuarios/{usuario_id}/reactivar")
def reactivar_usuario(
    usuario_id: int,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No tienes permisos de administrador")
    usuario = db.query(models.Usuario).filter(models.Usuario.id == usuario_id).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    usuario.esta_activo = True
    db.commit()
    return {"message": f"Usuario {usuario.email} reactivado", "usuario_id": usuario_id}

#AUTENTICACION =================================
@app.post("/token", response_model=schemas.Token)
async def login_for_access_token(
    # Esto acepta tanto JSON como form-data
    form_data: OAuth2PasswordRequestForm = Depends(),
    db: Session = Depends(get_db)
):
    """
    Endpoint para login:
    - Accepta application/x-www-form-urlencoded (para Swagger UI)
    - Accepta application/json (para tus clientes)
    """
    print("="*60)
    print("🔐 LOGIN ENDPOINT")
    print(f"Username recibido: {form_data.username}")
    print(f"Password recibida: {'*' * len(form_data.password)}")
    
    # Buscar usuario por email (username = email)
    user = db.query(models.Usuario).filter(
        models.Usuario.email == form_data.username
    ).first()
    
    if not user:
        print("❌ Usuario no encontrado")
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Email o contraseña incorrectos",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    # Verificar contraseña
    from app.auth import verify_password
    if not verify_password(form_data.password, user.hashed_password):
        print("❌ Contraseña incorrecta")
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Email o contraseña incorrectos",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    # Crear token
    from app.auth import create_access_token, ACCESS_TOKEN_EXPIRE_MINUTES
    #from datetime import timedelta
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.email}, expires_delta=access_token_expires
    )
    
    print(f"✅ Login exitoso para: {user.email}")
    print("="*60)
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "expires_in": ACCESS_TOKEN_EXPIRE_MINUTES * 60
    }



#patch
# main.py - Agrega este endpoint ANTES de @app.get("/transacciones/{transaccion_id}")
from app.schemas import EstadoUpdate
@app.patch("/transacciones/{transaccion_id}", response_model=schemas.TransaccionResponse)
def actualizar_transaccion(
    transaccion_id: int,
    transaccion_update: schemas.TransaccionUpdate,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Actualizar una transacción (campos: tipo, estado, descripcion, monto, fecha, beneficiario, etc.)
    - Usuario dueño: solo si está PENDIENTE
    - Contador (misma empresa): puede editar PENDIENTE o CONFIRMADA
    - Admin: puede editar cualquier transacción
    """
    print("="*60)
    print(f"📝 PATCH - Actualizando transacción ID: {transaccion_id}")
    print(f"👤 Usuario: {current_user.email} (rol: {current_user.rol}, empresa_id: {current_user.empresa_id})")
    
    # Buscar la transacción (primero por usuario actual)
    transaccion = db.query(models.Transaccion).filter(
        models.Transaccion.id == transaccion_id,
        models.Transaccion.usuario_id == current_user.id
    ).first()
    
    # Si no es del usuario actual pero es admin o contador, buscar por empresa
    if not transaccion and current_user.rol in ["admin", "contador"]:
        transaccion = db.query(models.Transaccion).filter(
            models.Transaccion.id == transaccion_id,
            models.Transaccion.empresa_id == current_user.empresa_id
        ).first()
        if transaccion:
            print(f"⚠️ {current_user.rol} editando transacción de otro usuario (ID propietario: {transaccion.usuario_id})")
    
    if not transaccion:
        raise HTTPException(status_code=404, detail="Transacción no encontrada")
    
    # Verificar permisos según estado
    if transaccion.usuario_id == current_user.id:
        # Dueño: solo puede editar si está PENDIENTE
        if transaccion.estado == "CONFIRMADA" and current_user.rol != "admin":
            raise HTTPException(
                status_code=403,
                detail="No puedes editar una transacción confirmada. Contacta al contador o administrador."
            )
    elif current_user.rol not in ["admin", "contador"]:
        # Ni dueño ni admin/contador
        raise HTTPException(
            status_code=403,
            detail="No tienes permisos para editar esta transacción"
        )
    
    # Obtener solo los campos enviados (excluir None)
    update_data = transaccion_update.dict(exclude_unset=True)
    print(f"📦 Datos recibidos para actualizar: {update_data}")
    
    # Lista de campos permitidos (evitar actualizar campos sensibles como id, usuario_id, empresa_id)
    campos_permitidos = [
        "tipo", "estado", "descripcion", "etiquetas",
        "monto", "fecha", "banco", "beneficiario", "ordenante", "moneda"
    ]
    
    # Actualizar campo por campo
    for key, value in update_data.items():
        if key in campos_permitidos and value is not None:
            setattr(transaccion, key, value)
            print(f"   ✅ '{key}' actualizado: {value}")
        elif key not in campos_permitidos:
            print(f"   ⚠️ Campo '{key}' no permitido, ignorado")
    
    # Si el estado cambió a CONFIRMADA, registrar fecha_confirmacion
    if update_data.get('estado') == "CONFIRMADA" and transaccion.estado != "CONFIRMADA":
        from datetime import datetime
        transaccion.fecha_confirmacion = datetime.now()
        print(f"   📅 Fecha confirmación registrada: {transaccion.fecha_confirmacion}")
    
    # Si el estado cambió a PENDIENTE, limpiar fecha_confirmacion
    if update_data.get('estado') == "PENDIENTE" and transaccion.estado == "CONFIRMADA":
        transaccion.fecha_confirmacion = None
        print(f"   🗑️ Fecha confirmación eliminada")
    
    db.commit()
    db.refresh(transaccion)
    print("✅ Cambios guardados correctamente")
    print("="*60)
    
    return transaccion


# Mantén también tu endpoint JSON original si lo necesitas
@app.post("/token/json", response_model=schemas.Token)
def login_json(
    login_data: schemas.UserLogin,
    db: Session = Depends(get_db)
):
    """Endpoint alternativo que acepta JSON"""
    user = auth.authenticate_user(db, login_data.email, login_data.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Email o contraseña incorrectos",
        )
    
    access_token_expires = timedelta(minutes=auth.ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = auth.create_access_token(
        data={"sub": user.email}, expires_delta=access_token_expires
    )
    print(access_token)
    return {"access_token": access_token, "token_type": "bearer"}

@app.post("/procesar-texto")
def procesar_texto_transaccion(texto: str,
    db: Session = Depends(get_db)
):
    """Endpoint público para parsear texto sin guardar en DB"""
    try:
        datos = utils.parse_transaccion_texto(texto)
        
        print(datos.values())
        
        return {
            "success": True,
            "data": datos,
            "texto_enmascarado": {
                "beneficiario": utils.enmascarar_numero_cuenta(datos["beneficiario"]),
                "ordenante": utils.enmascarar_numero_cuenta(datos["ordenante"])
            }
        }
            
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al procesar: {str(e)}")
    
    
# Endpoints protegidos
#para movil 

@app.post("/transacciones/recibir", response_model=schemas.TransaccionResponse)
async def recibir_transaccion(
    texto: str = Form(...),
    tipo: str = Form(...),
    token: str = Form(...),
    descripcion: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """
    Endpoint para apps móviles - Versión con normalización de tipo
    """
    try:
        print("="*70)
        print("📱 RECIBIDA TRANSACCIÓN")
        print("="*70)
        
        # 1. Autenticar
        from app.auth import get_current_user_from_token
        current_user = await get_current_user_from_token(token, db)
        print(f"👤 Usuario: {current_user.email} (Empresa ID: {current_user.empresa_id})")
        
        # 2. NORMALIZAR TIPO
        tipo_normalizado = tipo.lower().strip()
        if tipo_normalizado not in ['enviada', 'recibida']:
            raise HTTPException(status_code=400, detail=f"Tipo inválido")
        
        # 3. Normalizar texto
        from app.utils import normalizar_texto_transferencia, parse_transaccion_texto, normalizar_fecha
        texto_normalizado = normalizar_texto_transferencia(texto)
        
        # 4. Parsear
        datos = parse_transaccion_texto(texto_normalizado)
        
        if datos == "Mensaje Incorrecto":
            raise HTTPException(status_code=500, detail="El texto del mensaje no pudo ser procesado")
        
        print(f"📊 Datos parseados: {datos}")
        
        # 5. NORMALIZAR FECHA
        if datos.get('fecha'):
            datos['fecha'] = normalizar_fecha(datos['fecha'])
        
        # 6. Verificar campos requeridos
        campos_requeridos = ['fecha', 'beneficiario', 'ordenante', 'monto', 'numero_transaccion']
        for campo in campos_requeridos:
            if not datos.get(campo):
                if campo == 'beneficiario':
                    match = re.search(r'Beneficiario:\s*(\S+)', texto)
                    if match:
                        datos[campo] = match.group(1)
                        continue
                raise HTTPException(status_code=400, detail=f"Campo requerido no encontrado: {campo}")
        
        # ========== VALIDACIÓN DE DUPLICADOS (CORREGIDA) ==========
        numero_transaccion = datos['numero_transaccion']
        print(f"🔍 Verificando si la transacción {numero_transaccion} ya existe...")
        
        # 1. Verificar si existe para el MISMO usuario
        transaccion_mismo_usuario = db.query(models.Transaccion).filter(
            models.Transaccion.numero_transaccion == numero_transaccion,
            models.Transaccion.usuario_id == current_user.id
        ).first()
        
        if transaccion_mismo_usuario:
            print(f"⚠️ Transacción ya existe para este usuario (ID: {transaccion_mismo_usuario.id})")
            #return transaccion_mismo_usuario 

            raise HTTPException(
                status_code=409,
                detail=f"La transacción {numero_transaccion} ya está registrada para este usuario"
            )
        
        # 2. Verificar si existe para OTRO usuario de la misma empresa
        transaccion_otro_usuario = db.query(models.Transaccion).filter(
            models.Transaccion.numero_transaccion == numero_transaccion,
            models.Transaccion.empresa_id == current_user.empresa_id,
            models.Transaccion.usuario_id != current_user.id
        ).first()
        
        if transaccion_otro_usuario:
            print(f"⚠️ Transacción ya existe para otro usuario (ID: {transaccion_otro_usuario.id})")
            
            raise HTTPException(
                status_code=409,
                detail=f"La transacción {numero_transaccion} ya fue registrada por otro usuario"
            )
            
        
        print(f"✅ Transacción no existe, procediendo a crear...")
        # ========== FIN VALIDACIÓN ==========
        
        # 7. Crear transacción
        nueva_transaccion = models.Transaccion(
            usuario_id=current_user.id,
            empresa_id=current_user.empresa_id,
            banco=datos.get('banco', 'Banco Popular'),
            fecha=datos['fecha'],
            beneficiario=datos['beneficiario'],
            ordenante=datos['ordenante'],
            monto=float(datos['monto']),
            moneda=datos.get('moneda', 'CUP'),
            numero_transaccion=numero_transaccion,
            tipo=tipo_normalizado,
            estado=models.EstadoTransaccion.PENDIENTE,
            descripcion=descripcion or datos.get('concepto', ''),
            etiquetas="[]"
        )
        
        db.add(nueva_transaccion)
        db.commit()
        db.refresh(nueva_transaccion)
        
        print("✅ Transacción guardada")
        return nueva_transaccion
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Mensaje Incorrecto: {str(e)}")



@app.post("/transacciones/parsear", response_model=schemas.TransaccionResponse)
def crear_transaccion_desde_texto(
    transaccion_req: schemas.TransaccionParseRequest,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """Crear transacción a partir de texto bancario"""
    try:
        print("="*60)
        print("📥 Recibida petición de parseo")
        print(f"Texto original: {repr(transaccion_req.texto)}")
        print(f"Tipo: {transaccion_req.tipo}")
        print(f"Descripción: {transaccion_req.descripcion}")
        
        # PASO 1: Verificar que el texto no esté vacío
        if not transaccion_req.texto or not transaccion_req.texto.strip():
            raise HTTPException(
                status_code=400,
                detail="El texto de la transacción no puede estar vacío"
            )
        
        # PASO 2: Limpiar texto (eliminar espacios extras, tabs, etc.)
        texto_limpio = utils.limpiar_texto(transaccion_req.texto)
        print(f"Texto limpio: {repr(texto_limpio)}")
        
        # PASO 3: Parsear el texto
        try:
            datos = utils.parse_transaccion_texto(texto_limpio)
            print(f"Datos parseados: {datos}")
        except Exception as parse_error:
            print(f"❌ Error en parseo: {str(parse_error)}")
            raise HTTPException(
                status_code=400,
                detail=f"No se pudo parsear el texto. Asegúrate que contenga: fecha, beneficiario, ordenante, monto. Error: {str(parse_error)}"
            )
        # PASO 4: Verificar si ya existe para este usuario
        existente = db.query(models.Transaccion).filter(
            models.Transaccion.numero_transaccion == datos["numero_transaccion"],
            models.Transaccion.usuario_id == current_user.id
        ).first()
        
        # verificar si existe para otro usuario
        existe_otro_usuario = db.query(models.Usuario.transacciones).filter(
            models.Usuario.transacciones.numero_transaccion == datos["numero_transaccion"],
            
            
        )
        '''if existe_otro_usuario:
            existente = existe_otro_usuario '''
        
        
        
        if existente:
            print(f"⚠️ Transacción duplicada: {datos['numero_transaccion']}")
            # Opcional: Devolver la existente en lugar de error
            raise HTTPException(
                status_code=400, 
                detail=f"La transacción {datos['numero_transaccion']} ya existe (ID: {existente.id})"
            )
        
        
        
        
        
        # PASO 5: Validar que se obtuvieron los datos mínimos necesarios
        campos_requeridos = ['fecha', 'beneficiario', 'ordenante', 'monto', 'numero_transaccion']
        campos_faltantes = [campo for campo in campos_requeridos if not datos.get(campo)]
        
        if campos_faltantes:
            raise HTTPException(
                status_code=400,
                detail=f"No se pudieron extraer los siguientes campos del texto: {', '.join(campos_faltantes)}"
            )
        
        
        # PASO 6: Crear transacción
        try:
            transaccion = crud.parse_and_create_transaccion(
                db=db,
                texto=texto_limpio,
                tipo=transaccion_req.tipo,
                usuario_id=current_user.id,
                descripcion=transaccion_req.descripcion
            )
        except Exception as create_error:
            print(f"❌ Error al crear en BD: {str(create_error)}")
            raise HTTPException(
                status_code=500,
                detail=f"Error al guardar la transacción: {str(create_error)}"
            )
        
        print("✅ Transacción creada exitosamente")
        print(f"ID: {transaccion.id}")
        print(f"Número: {transaccion.numero_transaccion}")
        print("="*60)
        
        return transaccion
        
    except HTTPException:
        # Re-lanzar excepciones HTTP para que FastAPI las maneje
        raise
    except Exception as e:
        print(f"❌ Error inesperado: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500, 
            detail=f"Error interno del servidor: {str(e)}"
        )
@app.post("/transacciones/", response_model=schemas.TransaccionResponse)
def crear_transaccion(
    transaccion: schemas.TransaccionCreate,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    
    """Crear transacción manualmente"""
    print(transaccion)
    print("tipo de dato ",type(transaccion))
    return crud.create_transaccion(db=db, transaccion=transaccion, usuario_id=current_user.id)

def get_transacciones_empresa(db: Session, empresa_id: int, **filtros):
    """Obtener transacciones de una empresa"""
    query = db.query(models.Transaccion).filter(
        models.Transaccion.empresa_id == empresa_id
    )
    # ... aplicar filtros ...
    return query.all()






@app.get("/transacciones/")
def listar_transacciones(
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    # Usuario ve transacciones de su empresa
    transacciones = get_transacciones_empresa(db, current_user.empresa_id)
    return transacciones





@app.get("/transacciones/buscar", response_model=List[schemas.TransaccionResponse])
def buscar_transacciones(
    q: str = Query(..., min_length=1, description="Término de búsqueda"),
    tipo: Optional[str] = Query(None, description="Filtrar por tipo: enviada o recibida"),
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Buscar transacciones del usuario actual.
    IMPORTANTE: Solo se pueden buscar transacciones propias.
    """
    print(f"🔍 Búsqueda: '{q}', tipo: '{tipo}'")
    print(f"👤 Usuario: {current_user.email} (ID: {current_user.id}, rol: {current_user.rol})")
    
    # Verificar si el usuario tiene transacciones en general
    total_transacciones = db.query(models.Transaccion).filter(
        models.Transaccion.usuario_id == current_user.id
    ).count()
    
    print(f"📊 Total de transacciones del usuario: {total_transacciones}")
    
    if total_transacciones == 0:
        print(f"⚠️ El usuario {current_user.email} no tiene transacciones registradas")
        # Devolver lista vacía con información en el header (opcional)
        # O simplemente devolver lista vacía
        return []
    
    # Construir consulta base: SOLO transacciones del usuario actual
    query = db.query(models.Transaccion).filter(
        models.Transaccion.usuario_id == current_user.id
    )
    
    # Aplicar filtro de tipo si existe
    if tipo:
        tipo_normalizado = tipo.lower().strip()
        if tipo_normalizado in ['enviada', 'recibida']:
            query = query.filter(models.Transaccion.tipo == tipo_normalizado)
            print(f"   Filtro por tipo: {tipo_normalizado}")
    
    # Aplicar búsqueda por texto en varios campos
    search_pattern = f"%{q}%"
    query = query.filter(
        (models.Transaccion.numero_transaccion.ilike(search_pattern)) |
        (models.Transaccion.beneficiario.ilike(search_pattern)) |
        (models.Transaccion.ordenante.ilike(search_pattern)) |
        (models.Transaccion.banco.ilike(search_pattern)) |
        (models.Transaccion.descripcion.ilike(search_pattern))
    )
    
    # Ordenar y limitar
    transacciones = query.order_by(
        models.Transaccion.fecha.desc()
    ).limit(50).all()
    
    print(f"📊 Transacciones encontradas con el filtro: {len(transacciones)}")
    
    if len(transacciones) == 0 and total_transacciones > 0:
        print(f"⚠️ No se encontraron coincidencias para '{q}' en las transacciones del usuario")
    
    # Normalizar datos para evitar errores de validación
    for t in transacciones:
        if not t.banco or len(t.banco) < 2:
            t.banco = "Banco no especificado"
        if not t.beneficiario or len(t.beneficiario) < 4:
            t.beneficiario = "XXXX"
        if not t.ordenante or len(t.ordenante) < 4:
            t.ordenante = "XXXX"
        if t.descripcion is None:
            t.descripcion = ""
    
    return transacciones







@app.get("/transacciones/{transaccion_id}", response_model=schemas.TransaccionResponse)
def obtener_transaccion(
    transaccion_id: int,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """Obtener una transacción específica"""
    transaccion = crud.get_transaccion(db, transaccion_id=transaccion_id, usuario_id=current_user.id)
    if transaccion is None:
        raise HTTPException(status_code=404, detail="Transacción no encontrada")
    return transaccion


# main.py - Agregar este endpoint
@app.get("/eliminar-transaccion", response_class=HTMLResponse)
async def eliminar_transaccion_page(request: Request):
    """Página para eliminar transacciones"""
    return templates.TemplateResponse("eliminar_transaccion.html", {"request": request})
@app.delete("/transacciones/{transaccion_id}")
def eliminar_transaccion(
    transaccion_id: int,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Eliminar una transacción (solo el usuario propietario o admin puede hacerlo)
    """
    print("="*60)
    print(f"🗑️ Eliminando transacción ID: {transaccion_id}")
    print(f"👤 Usuario: {current_user.email} (rol: {current_user.rol})")
    
    # Buscar la transacción
    transaccion = db.query(models.Transaccion).filter(
        models.Transaccion.id == transaccion_id,
        models.Transaccion.usuario_id == current_user.id
    ).first()
    
    # Si es admin, puede eliminar cualquier transacción
    if not transaccion and current_user.rol == "admin":
        transaccion = db.query(models.Transaccion).filter(
            models.Transaccion.id == transaccion_id
        ).first()
        if transaccion:
            print(f"⚠️ Admin eliminando transacción de otro usuario (ID: {transaccion.usuario_id})")
    
    if not transaccion:
        raise HTTPException(status_code=404, detail="Transacción no encontrada")
    
    # Guardar información para log
    info = f"{transaccion.fecha} - {transaccion.numero_transaccion} - {transaccion.monto} {transaccion.moneda}"
    
    # Eliminar
    db.delete(transaccion)
    db.commit()
    
    print(f"✅ Transacción eliminada: {info}")
    print("="*60)
    
    return {"message": "Transacción eliminada correctamente", "id": transaccion_id, "info": info}


@app.put("/transacciones/{transaccion_id}", response_model=schemas.TransaccionResponse)
def actualizar_transaccion_completa(
    transaccion_id: int,
    transaccion_actualizada: schemas.TransaccionCreate,  # Usa Create que tiene todos los campos
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """Actualizar transacción completa (PUT)"""
    
    print("="*60)
    print(f"📝 ACTUALIZANDO TRANSACCIÓN ID: {transaccion_id}")
    print(f"Datos recibidos: {transaccion_actualizada.model_dump()}")
    
    # Verificar que existe y pertenece al usuario
    db_transaccion = db.query(models.Transaccion).filter(
        models.Transaccion.id == transaccion_id,
        models.Transaccion.usuario_id == current_user.id
    ).first()
    
    if not db_transaccion:
        raise HTTPException(status_code=404, detail="Transacción no encontrada")
    
    update_data = transaccion_actualizada.model_dump()
    
    if 'etiquetas' in update_data:
        import json
        
        update_data['etiquetas'] = json.dumps(update_data['etiquetas'])
        del update_data['etiquetas']
        print(f" campo etiquetas borrado {update_data}")
    
  
    
    # Actualizar campo por campo
    '''for key, value in transaccion_actualizada.dict().items():
        setattr(db_transaccion, key, value)
        print(f"   Campo '{key}' actualizado a: {value}")'''
        
    for key, value in update_data.items():
        if value is not None:
            setattr(db_transaccion, key, value)
            print(f"   Campo '{key}' actualizado a: {value}")

    db.commit()
    db.refresh(db_transaccion)
    print("✅ Cambios guardados en BD")
    print("="*60)
    
    return db_transaccion

#end point confirmat transacciones 
@app.get("/confirmar", response_class=HTMLResponse)
async def confirmar_transacciones_page(request: Request):
    """Página para que el contador confirme transferencias"""
    return templates.TemplateResponse("confirmar_transacciones.html", {"request": request})





@app.post("/transacciones/{transaccion_id}/confirmar", response_model=schemas.TransaccionResponse)
def confirmar_transaccion(
    transaccion_id: int,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """Confirmar una transacción recibida"""
    transaccion = crud.confirmar_transaccion(
        db=db,
        transaccion_id=transaccion_id,
        usuario_id=current_user.id
    )
    if transaccion is None:
        raise HTTPException(status_code=404, detail="Transacción no encontrada")
    return transaccion
#subir imagenes 

@app.get("/subir-imagen", response_class=HTMLResponse)
async def subir_imagen_page(request: Request):
    """
    Página para subir capturas de pantalla de transferencias
    """
    return templates.TemplateResponse("subir_imagen.html", {"request": request})





from fastapi import File, UploadFile
import io
@app.post("/transacciones/from-image", response_model=schemas.TransaccionResponse)
async def crear_transaccion_desde_imagen(
    imagen: UploadFile = File(...),
    tipo: str = Form(...),
    descripcion: Optional[str] = Form(None),
    token: str = Form(...),
    db: Session = Depends(get_db)
):
    """
    Crea una transacción a partir de una imagen (captura de pantalla)
    """
    print("="*70)
    print("📸 PROCESANDO IMAGEN DE TRANSFERENCIA")
    print("="*70)

    # 1. Autenticación
    from app.auth import get_current_user_from_token
    try:
        current_user = await get_current_user_from_token(token, db)
    except Exception as auth_error:
        raise HTTPException(status_code=401, detail=f"No autorizado: {str(auth_error)}")
    
    print(f"👤 Usuario: {current_user.email} (ID: {current_user.id}, empresa_id: {current_user.empresa_id})")

    # 2. Validar tipo de archivo
    if not imagen.content_type or not imagen.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="El archivo debe ser una imagen (JPG, PNG, etc.)")

    # 3. Leer imagen
    try:
        imagen_bytes = await imagen.read()
        if len(imagen_bytes) == 0:
            raise HTTPException(status_code=400, detail="La imagen está vacía")
        print(f"📦 Tamaño de imagen: {len(imagen_bytes)} bytes")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer la imagen: {str(e)}")

    # 4. OCR - Extraer texto
    try:
        from app.utils import extraer_texto_de_imagen
        texto_extraido = extraer_texto_de_imagen(imagen_bytes)
    except Exception as ocr_error:
        print(f"❌ Error en OCR: {ocr_error}")
        raise HTTPException(status_code=400, detail="No se pudo procesar la imagen con OCR. Asegúrate de que sea clara y tenga texto visible.")

    if not texto_extraido or len(texto_extraido.strip()) < 10:
        raise HTTPException(status_code=400, detail="No se pudo extraer texto suficiente de la imagen. Prueba con una captura más nítida.")

    print(f"📝 Texto extraído:\n{texto_extraido}")

    # 5. Extraer datos de transferencia del texto
    try:
        from app.utils import extraer_datos_transferencia_de_texto_ocr
        datos = extraer_datos_transferencia_de_texto_ocr(texto_extraido)
    except Exception as extraction_error:
        print(f"❌ Error al interpretar datos: {extraction_error}")
        raise HTTPException(status_code=400, detail="No se pudieron interpretar los datos de la transferencia desde el texto extraído.")

    # Validar que se haya extraído al menos algo útil
    if not any([datos.get("monto"), datos.get("beneficiario"), datos.get("numero_transaccion")]):
        raise HTTPException(status_code=400, detail="No se encontraron datos válidos de transferencia (monto, beneficiario o número de transacción).")

    print("📊 Datos extraídos:")
    for key, value in datos.items():
        print(f"   {key}: {value}")

    # 6. Normalizar tipo
    tipo_normalizado = tipo.lower().strip()
    if tipo_normalizado not in ['enviada', 'recibida']:
        raise HTTPException(status_code=400, detail="El tipo debe ser 'enviada' o 'recibida'")

    # 7. Generar/validar número de transacción
    numero_transaccion = datos.get("numero_transaccion")
    if not numero_transaccion or len(numero_transaccion) < 4:
        from datetime import datetime
        numero_transaccion = f"OCR{datetime.now().strftime('%Y%m%d%H%M%S')}"
        print(f"⚠️ Número de transacción no encontrado, se generó automático: {numero_transaccion}")
    else:
        # Limpiar caracteres no deseados
        import re
        numero_transaccion = re.sub(r'[^\w\d-]', '', numero_transaccion).upper()

    # 8. Verificar duplicados en la misma empresa
    print(f"🔍 Verificando si la transacción {numero_transaccion} ya existe...")
    transaccion_existente = db.query(models.Transaccion).filter(
        models.Transaccion.numero_transaccion == numero_transaccion,
        models.Transaccion.empresa_id == current_user.empresa_id
    ).first()

    if transaccion_existente:
        print(f"✅ Transacción ya existe (ID: {transaccion_existente.id})")
        # Normalizar campos mínimos si están vacíos (para evitar None)
        if transaccion_existente.empresa_id is None:
            transaccion_existente.empresa_id = current_user.empresa_id
            db.commit()
        if not transaccion_existente.banco:
            transaccion_existente.banco = "Banco no especificado"
        if not transaccion_existente.beneficiario:
            transaccion_existente.beneficiario = "XXXX"
        if not transaccion_existente.ordenante:
            transaccion_existente.ordenante = "XXXX"
        if transaccion_existente.descripcion is None:
            transaccion_existente.descripcion = ""
        db.commit()
        return transaccion_existente

    print("✅ Transacción no existe, procediendo a crear...")

    # 9. Crear nueva transacción
    try:
        nueva_transaccion = models.Transaccion(
            usuario_id=current_user.id,
            empresa_id=current_user.empresa_id,
            banco=datos.get('banco') or "Banco no especificado",
            fecha=datos.get('fecha'),
            beneficiario=datos.get('beneficiario') or "XXXX-xxxx-xxxx-XXXX",
            ordenante=datos.get('ordenante') or "XXXX-xxxx-xxxx-XXXX",
            monto=float(datos.get('monto', 0.00)),
            moneda=datos.get('moneda', 'CUP'),
            numero_transaccion=numero_transaccion,
            tipo=tipo_normalizado,
            estado=models.EstadoTransaccion.PENDIENTE,
            descripcion=descripcion or datos.get('concepto') or f"OCR: {texto_extraido[:100]}",
            etiquetas="[]"
        )
        db.add(nueva_transaccion)
        db.commit()
        db.refresh(nueva_transaccion)
        print(f"✅ Transacción creada desde imagen. ID: {nueva_transaccion.id}")
        return nueva_transaccion

    except Exception as create_error:
        db.rollback()
        print(f"❌ Error al guardar en BD: {create_error}")
        raise HTTPException(status_code=500, detail=f"Error al guardar la transacción: {str(create_error)}")

# Reportes
@app.get("/reportes/resumen", response_model=schemas.ResumenTransacciones)
def obtener_resumen(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """Obtener resumen de transacciones"""
    return crud.get_resumen_transacciones(
        db=db,
        usuario_id=current_user.id,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin
    )

@app.get("/reportes/enviadas")
def reporte_transacciones_enviadas(
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """Reporte detallado de transacciones enviadas"""
    transacciones = crud.get_transacciones_usuario(
        db=db,
        usuario_id=current_user.id,
        tipo=schemas.TipoTransaccion.ENVIADA
    )
    
    total = sum(t.monto for t in transacciones)
    return {
        "total_transacciones": len(transacciones),
        "total_monto": total,
        "transacciones": transacciones
    }

@app.get("/reportes/recibidas")
def reporte_transacciones_recibidas(
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """Reporte detallado de transacciones recibidas"""
    transacciones = crud.get_transacciones_usuario(
        db=db,
        usuario_id=current_user.id,
        tipo=schemas.TipoTransaccion.RECIBIDA
    )
    
    pendientes = [t for t in transacciones if t.estado == schemas.EstadoTransaccion.PENDIENTE]
    confirmadas = [t for t in transacciones if t.estado == schemas.EstadoTransaccion.CONFIRMADA]
    
    total = sum(t.monto for t in transacciones)
    return {
        "total_transacciones": len(transacciones),
        "pendientes": len(pendientes),
        "confirmadas": len(confirmadas),
        "total_monto": total,
        "transacciones": transacciones
    }

# Etiquetas
@app.post("/etiquetas/", response_model=schemas.EtiquetaResponse)
def crear_etiqueta(
    etiqueta: schemas.EtiquetaCreate,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """Crear una nueva etiqueta"""
    return crud.create_etiqueta(db=db, etiqueta=etiqueta, usuario_id=current_user.id)

@app.get("/etiquetas/", response_model=List[schemas.EtiquetaResponse])
def listar_etiquetas(
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """Listar etiquetas del usuario"""
    return crud.get_etiquetas_usuario(db=db, usuario_id=current_user.id)

# Perfil de usuario
@app.get("/perfil", response_model=schemas.UserResponse)
def obtener_perfil(
    current_user: models.Usuario = Depends(auth.get_current_active_user)
):
    """Obtener perfil del usuario actual"""
    return current_user

# Health check
@app.get("/health")
def health_check():
    return {"status": "healthy", "service": "gestor-bancario"}

@app.get("/transacciones/estadisticas")
def obtener_estadisticas(
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """Obtener estadísticas detalladas del usuario"""
    return crud.get_estadisticas_usuario(db, current_user.id)

@app.get("/transacciones/por-mes/{año}/{mes}")
def transacciones_por_mes(
    año: int,
    mes: int,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """Obtener transacciones de un mes específico"""
    return crud.get_transacciones_por_mes(db, current_user.id, año, mes)




# prueba para debuguear
@app.post("/transacciones/recibir", response_model=schemas.TransaccionResponse)
async def recibir_transaccion_debug(
    texto: str = Form(...),
    tipo: schemas.TipoTransaccion = Form(...),
    token: str = Form(...),
    descripcion: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """
    Versión con debugging del endpoint
    """
    print("="*80)
    print("🔍 DEBUG - RECIBIR TRANSACCIÓN")
    print("="*80)
    print(f"Token recibido: {token[:30]}...")
    print(f"Tipo: {tipo}")
    print(f"Texto: {repr(texto)[:100]}...")
    
    from app.auth import SECRET_KEY, ALGORITHM
    from jose import jwt
    
    try:
        # Intentar decodificar el token manualmente
        print("\n📦 Decodificando token...")
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        print(f"✅ Token decodificado:")
        print(f"   sub: {payload.get('sub')}")
        print(f"   exp: {datetime.fromtimestamp(payload.get('exp'))}")
        print(f"   iat: {datetime.fromtimestamp(payload.get('iat'))}")
        
        # Buscar usuario
        email = payload.get('sub')
        user = db.query(models.Usuario).filter(models.Usuario.email == email).first()
        
        if user:
            print(f"✅ Usuario encontrado en BD: {user.email}")
            print(f"   ID: {user.id}")
        else:
            print(f"❌ Usuario NO encontrado en BD: {email}")
            raise HTTPException(status_code=401, detail="Usuario no existe")
        
    except jwt.ExpiredSignatureError:
        print("❌ Token expirado")
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.JWTError as e:
        print(f"❌ Error decodificando token: {str(e)}")
        raise HTTPException(status_code=401, detail=f"Token inválido: {str(e)}")
    
    print("="*80)
    
    # Continuar con la creación de la transacción...
    return {"mensaje": "Debug completado", "usuario": user.email}


# routers

from fastapi import FastAPI, Depends, HTTPException, status, Form, Request
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse

from typing import List, Optional
from datetime import timedelta, datetime
import os

# Configurar templates
templates = Jinja2Templates(directory="templates")

# ... (tu código existente) ...


@app.get("/test", response_class=HTMLResponse)
async def test_page(request: Request):
    """
    Página de prueba para el endpoint de transferencias
    """
    return templates.TemplateResponse("login_simple.html", {"request": request})

@app.get("/capturar", response_class=HTMLResponse)
async def capturar_page(request: Request):
    """Página para capturar transferencias (principal para usuarios)"""
    return templates.TemplateResponse("capturar.html", {"request": request})


@app.get("/transacciones/enviadas/por-fecha", response_model=List[schemas.TransaccionResponse])
async def listar_transacciones_enviadas_por_fecha(
    fecha_inicio: str = Query(..., description="Fecha inicial en formato DD/MM/YYYY (ej: 01/03/2026)"),
    fecha_fin: str = Query(..., description="Fecha final en formato DD/MM/YYYY (ej: 31/03/2026)"),
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    📤 Obtiene todas las transferencias ENVIADAS de un usuario en un rango de fechas
    
    - Requiere autenticación (token JWT)
    - Las fechas deben estar en formato DD/MM/YYYY
    - Retorna lista de transacciones ordenadas por fecha descendente
    """
    try:
        print("="*70)
        print(f"📤 LISTAR TRANSACCIONES ENVIADAS POR FECHA")
        print("="*70)
        print(f"👤 Usuario: {current_user.email} (ID: {current_user.id})")
        print(f"📅 Rango: {fecha_inicio} - {fecha_fin}")
        
        # Validar formato de fechas
        from app.utils import normalizar_fecha
        try:
            fecha_inicio_norm = normalizar_fecha(fecha_inicio)
            fecha_fin_norm = normalizar_fecha(fecha_fin)
        except Exception as e:
            raise HTTPException(
                status_code=400,
                detail=f"Formato de fecha inválido: {str(e)}"
            )
        
        # Obtener transacciones
        transacciones = crud.get_transacciones_enviadas_por_fecha(
            db=db,
            usuario_id=current_user.id,
            fecha_inicio=fecha_inicio_norm,
            fecha_fin=fecha_fin_norm
        )
        
        print(f"✅ Se encontraron {len(transacciones)} transacciones")
        print("="*70)
        
        return transacciones
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")

@app.get("/transacciones/recibidas/por-fecha", response_model=List[schemas.TransaccionResponse])
async def listar_transacciones_recibidas_por_fecha(
    fecha_inicio: str = Query(..., description="Fecha inicial en formato DD/MM/YYYY"),
    fecha_fin: str = Query(..., description="Fecha final en formato DD/MM/YYYY"),
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    📥 Obtiene todas las transferencias RECIBIDAS de un usuario en un rango de fechas
    
    - Requiere autenticación (token JWT)
    - Las fechas deben estar en formato DD/MM/YYYY
    - Retorna lista de transacciones ordenadas por fecha descendente
    """
    try:
        transacciones = crud.get_transacciones_recibidas_por_fecha(
            db=db,
            usuario_id=current_user.id,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin
        )
        return transacciones
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/transacciones/enviadas/resumen")
async def resumen_transacciones_enviadas(
    fecha_inicio: str = Query(..., description="Fecha inicial DD/MM/YYYY"),
    fecha_fin: str = Query(..., description="Fecha final DD/MM/YYYY"),
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    📊 Obtiene un resumen de las transacciones enviadas en un rango de fechas
    """
    transacciones = crud.get_transacciones_enviadas_por_fecha(
        db=db,
        usuario_id=current_user.id,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin
    )
    
    if not transacciones:
        return {
            "total": 0,
            "monto_total": 0,
            "promedio": 0,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "mensaje": "No hay transacciones en el período"
        }
    
    total_monto = sum(t.monto for t in transacciones)
    
    return {
        "total_transacciones": len(transacciones),
        "monto_total": round(total_monto, 2),
        "promedio": round(total_monto / len(transacciones), 2),
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "transacciones": [
            {
                "id": t.id,
                "fecha": t.fecha,
                "monto": t.monto,
                "moneda": t.moneda,
                "beneficiario": t.beneficiario,
                "numero_transaccion": t.numero_transaccion
            }
            for t in transacciones[:10]  # Últimas 10
        ]
    }



# Configurar templates
#crear transacciones manualmente 
@app.get("/crear-transaccion", response_class=HTMLResponse)
async def crear_transaccion_manual(request: Request):
    """
    Página para crear transacciones manualmente
    """
    return templates.TemplateResponse("crear_transaccion.html", {"request": request})
#servir dashboard 
@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request):
    """
    Panel de usuario para consultar transacciones
    """
    return templates.TemplateResponse("dashboard.html", {"request": request})


@app.get("/dashboard_unificado", response_class=HTMLResponse)
async def dashboard(request: Request):
    """
    Panel de usuario para consultar transacciones
    """
    return templates.TemplateResponse("dashboard_unificado.html", {"request": request})





#EDITAR TRANSACCIONES 
# main.py - Agregar después de tus endpoints existentes

# main.py - Versión corregida del endpoint de búsqueda

@app.get("/transacciones/buscar", response_model=List[schemas.TransaccionResponse])
def buscar_transacciones_para_editar(
    q: str = Query(..., min_length=1, description="Término de búsqueda"),
    tipo: Optional[str] = Query(None, description="Filtrar por tipo: enviada o recibida"),
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Buscar transacciones del usuario actual para edición.
    """
    print(f"🔍 Búsqueda recibida - q: '{q}', tipo: '{tipo}'")  # Debug
    
    # Construir consulta
    query = db.query(models.Transaccion).filter(
        models.Transaccion.usuario_id == current_user.id
    )
    
    # Aplicar filtro de búsqueda si hay término
    if q and len(q) > 0:
        query = query.filter(
            (models.Transaccion.numero_transaccion.ilike(f"%{q}%")) |
            (models.Transaccion.beneficiario.ilike(f"%{q}%")) |
            (models.Transaccion.ordenante.ilike(f"%{q}%")) |
            (models.Transaccion.banco.ilike(f"%{q}%")) |
            (models.Transaccion.monto.ilike(f"%{q}%"))
        )
    
    # Aplicar filtro de tipo si existe
    if tipo:
        tipo_normalizado = tipo.lower().strip()
        if tipo_normalizado in ['enviada', 'recibida']:
            query = query.filter(models.Transaccion.tipo == tipo_normalizado)
    
    transacciones = query.order_by(models.Transaccion.fecha.desc()).limit(50).all()
    
    print(f"📊 Transacciones encontradas: {len(transacciones)}")  # Debug
    
    # Normalizar datos para respuesta (evitar errores de validación)
    for t in transacciones:
        if not t.banco or len(t.banco) < 2:
            t.banco = "Banco no especificado"
        if not t.beneficiario or len(t.beneficiario) < 4:
            t.beneficiario = "XXXX"
        if not t.ordenante or len(t.ordenante) < 4:
            t.ordenante = "XXXX"
    
    
    return transacciones




#SEERVIR EL HTML editar:transaccion
@app.get("/editar-transaccion", response_class=HTMLResponse)
async def editar_transaccion_page(request: Request):
    """Página para editar transacciones"""
    return templates.TemplateResponse("editar_transaccion.html", {"request": request})

# main.py - Agregar estos imports
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from fastapi.responses import StreamingResponse, FileResponse
import io
from datetime import datetime

# ========== EXPORTAR A EXCEL ==========
@app.get("/transacciones/exportar/excel")
def exportar_transacciones_excel(
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Exporta todas las transacciones del usuario a un archivo Excel (.xlsx)
    """
    print("="*60)
    print("📊 Exportando transacciones a Excel")
    print(f"👤 Usuario: {current_user.email}")
    
    # Obtener todas las transacciones del usuario
    transacciones = db.query(models.Transaccion).filter(
        models.Transaccion.usuario_id == current_user.id
    ).order_by(models.Transaccion.fecha.desc()).all()
    
    if not transacciones:
        raise HTTPException(status_code=404, detail="No hay transacciones para exportar")
    
    # Crear DataFrame con los datos
    data = []
    for t in transacciones:
        data.append({
            "ID": t.id,
            "Fecha": t.fecha,
            "N° Transacción": t.numero_transaccion,
            "Tipo": "Enviada" if t.tipo == "enviada" else "Recibida",
            "Beneficiario": t.beneficiario,
            "Ordenante": t.ordenante,
            "Monto": t.monto,
            "Moneda": t.moneda,
            "Banco": t.banco,
            "Estado": t.estado,
            "Descripción": t.descripcion or "",
            "Fecha Confirmación": t.fecha_confirmacion.strftime("%d/%m/%Y %H:%M") if t.fecha_confirmacion else "",
            "Fecha Procesamiento": t.fecha_procesamiento.strftime("%d/%m/%Y %H:%M") if t.fecha_procesamiento else ""
        })
    
    df = pd.DataFrame(data)
    
    # Crear archivo Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Transacciones', index=False)
        
        # Dar formato al archivo Excel
        worksheet = writer.sheets['Transacciones']
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Aplicar estilo a los encabezados
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    # Nombre del archivo con fecha
    fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"transacciones_{current_user.email}_{fecha_actual}.xlsx"
    
    print(f"✅ Excel generado: {filename}")
    print("="*60)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ========== EXPORTAR A PDF ==========
@app.get("/transacciones/exportar/pdf")
def exportar_transacciones_pdf(
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Exporta todas las transacciones del usuario a un archivo PDF
    """
    print("="*60)
    print("📊 Exportando transacciones a PDF")
    print(f"👤 Usuario: {current_user.email}")
    
    # Obtener todas las transacciones del usuario
    transacciones = db.query(models.Transaccion).filter(
        models.Transaccion.usuario_id == current_user.id
    ).order_by(models.Transaccion.fecha.desc()).all()
    
    if not transacciones:
        raise HTTPException(status_code=404, detail="No hay transacciones para exportar")
    
    # Crear buffer para PDF
    buffer = io.BytesIO()
    
    # Configurar documento (formato horizontal para más columnas)
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                            rightMargin=0.5*cm, leftMargin=0.5*cm,
                            topMargin=0.5*cm, bottomMargin=0.5*cm)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=1,  # Centro
        spaceAfter=20
    )
    
    # Contenido del PDF
    elements = []
    
    # Título
    titulo = f"Reporte de Transacciones - {current_user.nombre} {current_user.apellido}"
    elements.append(Paragraph(titulo, title_style))
    elements.append(Spacer(1, 20))
    
    # Fecha del reporte
    fecha_reporte = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    elements.append(Paragraph(f"Fecha de generación: {fecha_reporte}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Resumen estadístico
    total_enviadas = sum(t.monto for t in transacciones if t.tipo == "enviada")
    total_recibidas = sum(t.monto for t in transacciones if t.tipo == "recibida")
    balance = total_recibidas - total_enviadas
    
    resumen_data = [
        ["Total Transacciones", len(transacciones)],
        ["Total Enviadas", f"{total_enviadas:,.2f}"],
        ["Total Recibidas", f"{total_recibidas:,.2f}"],
        ["Balance", f"{balance:,.2f}"]
    ]
    
    resumen_tabla = Table(resumen_data, colWidths=[5*cm, 5*cm])
    resumen_tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(resumen_tabla)
    elements.append(Spacer(1, 20))
    
    # Datos de transacciones
    data = [["Fecha", "N° Transacción", "Tipo", "Beneficiario", "Monto", "Moneda", "Estado"]]
    
    for t in transacciones[:50]:  # Limitar a 50 registros para evitar PDF muy grande
        data.append([
            t.fecha or "",
            t.numero_transaccion or "",
            "Enviada" if t.tipo == "enviada" else "Recibida",
            t.beneficiario[:20] + "..." if len(t.beneficiario) > 20 else t.beneficiario,
            f"{t.monto:,.2f}",
            t.moneda,
            t.estado
        ])
    
    # Crear tabla
    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(tabla)
    
    # Pie de página
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"Total de transacciones: {len(transacciones)}", styles['Normal']))
    
    # Generar PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Nombre del archivo
    fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"transacciones_{current_user.email}_{fecha_actual}.pdf"
    
    print(f"✅ PDF generado: {filename}")
    print("="*60)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
#pagina principal
@app.get("/admin", response_class=HTMLResponse)
async def admin_page(request: Request):
    """Panel de administración"""
    return templates.TemplateResponse("admin1.html", {"request": request})

# main.py - Endpoint para crear usuarios (solo admin)

@app.post("/admin/usuarios", response_model=schemas.UserResponse)
def crear_usuario_por_admin(
    usuario: schemas.UserCreate,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """Crear nuevo usuario (solo administradores)"""
    print("usuario recibido: ",usuario)
    # Verificar que quien crea es admin
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No tienes permisos de administrador")
    
    # Verificar si el email ya existe
    db_user = crud.get_user_by_email(db, email=usuario.email)
    
    if db_user:
        raise HTTPException(status_code=400, detail="El email ya está registrado")
    
    # Validar que el rol sea válido
    rol_valido = usuario.rol if hasattr(usuario, 'rol') and usuario.rol in ["usuario", "contador", "admin"] else "usuario"
    if len(usuario.password) < 5:
        return "Pasword too Short "
    # Crear usuario con el rol seleccionado
    nuevo_usuario = models.Usuario(
        email=usuario.email,
        nombre=usuario.nombre,
        apellido=usuario.apellido,
        hashed_password=get_password_hash(usuario.password),
        esta_activo=True,
        rol=rol_valido,
        empresa_id=usuario.empresa_id
        )
    
    db.add(nuevo_usuario)
    db.commit()
    db.refresh(nuevo_usuario)
    
    return nuevo_usuario

@app.get("/admin/usuarios", response_model=List[schemas.UserResponse])
def listar_usuarios(
    skip: int = 0,
    limit: int = 100,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """Listar todos los usuarios (solo administradores)"""
    
    print(f"🔍 Usuario actual: {current_user.email}, rol: {current_user.rol}")
    
    if current_user.rol != "admin":
        print(f"❌ Acceso denegado: {current_user.rol} no es admin")
        raise HTTPException(status_code=403, detail="No tienes permisos de administrador")
    
    usuarios = db.query(models.Usuario).offset(skip).limit(limit).all()
    print(f"📊 Usuarios encontrados: {len(usuarios)}")
    
    return usuarios

# ========== ENDPOINTS PARA OPINIONES ==========

@app.post("/opiniones")
def crear_opinion(
    opinion: schemas.OpinionCreate,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """Crear una nueva opinión o sugerencia"""
    nueva_opinion = models.Opinion(
        usuario_id=current_user.id,
        usuario_nombre=f"{current_user.nombre} {current_user.apellido}",
        titulo=opinion.titulo,
        mensaje=opinion.mensaje,
        puntuacion=opinion.puntuacion
    )
    db.add(nueva_opinion)
    db.commit()
    db.refresh(nueva_opinion)
    return {"message": "Opinión enviada correctamente", "id": nueva_opinion.id}

@app.get("/opiniones")
def listar_opiniones(
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db)
):
    """Listar opiniones (público, solo lectura)"""
    opiniones = db.query(models.Opinion).order_by(
        models.Opinion.fecha.desc()
    ).offset(skip).limit(limit).all()
    return opiniones



@app.get("/admin/usuarios/{usuario_id}/transacciones", response_model=List[schemas.TransaccionResponse])
def ver_transacciones_usuario(
    usuario_id: int,
    skip: int = 0,
    limit: int = 100,
    current_user: models.Usuario = Depends(auth.get_current_active_user),
    db: Session = Depends(get_db)
):
    """Ver transacciones de un usuario específico (solo administradores)"""
    
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No tienes permisos de administrador")
    
    transacciones = db.query(models.Transaccion).filter(
        models.Transaccion.usuario_id == usuario_id
    ).order_by(models.Transaccion.fecha.desc()).offset(skip).limit(limit).all()
    
    return transacciones

@app.get("/clientes", response_class=HTMLResponse)
async def clientes_page(request: Request):
    """Página para ver información de clientes (datos en descripción)"""
    return templates.TemplateResponse("clientes.html", {"request": request})


